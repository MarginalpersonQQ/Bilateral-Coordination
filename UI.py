import tkinter as tk
from tkinter import filedialog, messagebox
import Action3_0
import threading
import os
import datetime
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

# 設定 Matplotlib 中文字體 (避免中文顯示為方塊)
plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'Taipei Sans TC Beta', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

file_var = None
video_fold_root_path = r"C:\Bilateral Coordination Record Video"
video_slots = None
root = None  # 全域 Tkinter 主視窗

excel_lock = threading.Lock()
start_time_str = None
amount_of_actions = 15

# 動作完成計數與分數暫存 (用於畫長條圖)
completed_actions = 0
action_results = {}
chart_lock = threading.Lock()

# 根據圖片定義每個動作對應的分數類型 (Index 0~14 對應 Action 1~15)
# 1: 順序, 2: 空間, 3: 時間
SCORE_TYPES = {
    0: ["順序", "空間", "時間"], # 動作1
    1: ["順序", "空間", "時間"], # 動作2
    2: ["順序", "空間", "時間"], # 動作3
    3: ["順序", "空間", "時間"], # 動作4
    4: ["順序"],
    5: ["順序", "空間", "時間"],
    6: ["順序", "空間", "時間"],
    7: ["順序", "空間", "時間"],
    8: ["順序"],
    9: ["順序", "空間", "時間"],
    10: ["順序", "空間", "時間"],
    11: ["順序", "空間", "時間"],
    12: ["順序", "空間", "時間"],
    13: ["順序", "空間", "時間"],
    14: ["順序", "空間", "時間"]
}


def init_excel_single_sheet():
    """建立單一工作表：橫向保留所有動作細項，最末端加上三個項目的平均值欄位"""
    excel_path = os.path.join(video_fold_root_path, "scores.xlsx")

    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Scores"

        row1 = ["影片名稱"]
        row2 = ["動作項目"]

        # 生成 15 個動作的細項欄位
        for idx in range(15):
            items = SCORE_TYPES[idx]
            for item in items:
                row1.append(f"Action {idx + 1}")
                row2.append(item)

        # 在最末端加上三個項目的統計欄位
        row1.extend(["總計平均", "總計平均", "總計平均"])
        row2.extend(["順序平均", "空間平均", "時間平均"])

        ws.append(row1)
        ws.append(row2)
        wb.save(excel_path)

    return excel_path

def save_scores_to_excel(video_name, index, scores_dict):
    excel_path = init_excel_single_sheet()

    with excel_lock:
        wb = load_workbook(excel_path)
        ws = wb.active

        # 計算前段 22 個細項所佔用的基本總欄位數 (2 欄起點 + 22 個細項 = 24)
        # 代表「順序平均」在第 24 欄、「空間平均」在第 25 欄、「時間平均」在第 26 欄
        base_col_count = 2 + sum(len(SCORE_TYPES[i]) for i in range(15))

        # 尋找影片名稱是否已存在於第 3 列之後
        target_row = None
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == video_name:
                target_row = row
                break

        # 若此影片第一次寫入，則建立新的一列並預填 "None"（包含最後三個平均欄位共 base_col_count + 3 欄）
        if target_row is None:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value=video_name)
            for c in range(2, base_col_count + 3):
                ws.cell(row=target_row, column=c, value="None")

        # 動態計算當前動作 (index) 在橫列中的起始細項欄位
        start_col = 2
        for i in range(index):
            start_col += len(SCORE_TYPES[i])

        # 寫入當前動作的細項分數
        mapping = SCORE_TYPES[index]
        for i, key in enumerate(mapping):
            val = scores_dict.get(key, "None")
            if val in ["None", "No Score"]:
                ws.cell(row=target_row, column=start_col + i, value="None")
            else:
                ws.cell(row=target_row, column=start_col + i, value=val)

        # 實時重新加總並計算該影片目前已獲得的分類平均值
        sums = {'順序': 0, '空間': 0, '時間': 0}
        counts = {'順序': 0, '空間': 0, '時間': 0}
        for action_idx, act_scores in action_results.items():
            for key, val in act_scores.items():
                if isinstance(val, (int, float)):
                    sums[key] += val
                    counts[key] += 1

        # 依序將 順序、空間、時間 的平均值分別填入末端欄位
        categories = ['順序', '空間', '時間']
        for offset, cat in enumerate(categories):
            target_col_idx = base_col_count + offset  # 第 24, 25, 26 欄
            if counts[cat] > 0:
                avg_val = sums[cat] / counts[cat]
                ws.cell(row=target_row, column=target_col_idx, value=round(avg_val, 2))  # 四捨五入至小數第二位
            else:
                ws.cell(row=target_row, column=target_col_idx, value="None")

        wb.save(excel_path)

def show_bar_chart():
    """計算平均值並顯示長條圖"""
    sums = {'順序': 0, '空間': 0, '時間': 0}
    counts = {'順序': 0, '空間': 0, '時間': 0}

    for action_idx, scores_dict in action_results.items():
        for key, val in scores_dict.items():
            if isinstance(val, (int, float)):
                sums[key] += val
                counts[key] += 1

    avgs = {}
    for key in sums:
        avgs[key] = sums[key] / counts[key] if counts[key] > 0 else 0

    labels = list(avgs.keys())
    values = list(avgs.values())

    plt.figure(figsize=(7, 5))
    bars = plt.bar(labels, values, color=['#4C72B0', '#55A868', '#C44E52'])
    plt.title(f"影片評分平均分析 - {file_var.get()}")
    plt.ylabel('平均分數')
    plt.ylim(0, max(max(values) * 1.2, 5))

    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width() / 2, yval, f"{yval:.2f}", ha='center', va='bottom')

    plt.show()

def update_grid(index, scores_dict):
    global video_slots

    mapping = SCORE_TYPES[index]
    for i, key in enumerate(mapping):
        val = scores_dict.get(key, "None")
        if val == "None" or val == "No Score":
            video_slots[index][i].config(text="None")
        else:
            video_slots[index][i].config(text=str(int(val)))

def worker(index, filename):
    global completed_actions, action_results

    video_path = file_var.get()
    full_path = os.path.join(video_fold_root_path, video_path, filename)

    scores_dict = {}
    mapping = SCORE_TYPES[index]

    if not os.path.isfile(full_path):
        for key in mapping:
            scores_dict[key] = "None"
    else:
        action_class_name = f"Action{index + 1}"
        action_class = getattr(Action3_0, action_class_name, None)

        if action_class is None:
            print(f"找不到對應類別: {action_class_name}")
            for key in mapping:
                scores_dict[key] = "None"
        else:
            result = action_class(full_path)
            result.main_func()

            raw_scores = result.score
            for i, key in enumerate(mapping):
                if i < len(raw_scores):
                    scores_dict[key] = raw_scores[i]
                else:
                    scores_dict[key] = "None"

    update_grid(index, scores_dict)
    save_scores_to_excel(video_path, index, scores_dict)

    with chart_lock:
        action_results[index] = scores_dict
        completed_actions += 1

        if completed_actions == amount_of_actions:
            root.after(0, show_bar_chart)

def run_all_actions(video_path):
    file_exten = ".mp4"
    filenames = [f"{i:02}{file_exten}" for i in range(1, 16)]

    for i, name in enumerate(filenames):
        threading.Thread(target=worker, args=(i, name), daemon=True).start()

def clear_all_slots():
    global video_slots
    if video_slots is None:
        return
    for slot_labels in video_slots:
        for label in slot_labels:
            label.config(text="")

def start_action():
    global start_time_str, completed_actions, action_results
    video_path = file_var.get()

    start_time_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with chart_lock:
        completed_actions = 0
        action_results.clear()

    clear_all_slots()
    messagebox.showinfo("開始判斷", f"開始處理影片：{video_path}")
    run_all_actions(video_path)

def get_video_files():
    if not os.path.exists(video_fold_root_path):
        return []
    return [f for f in os.listdir(video_fold_root_path)]

def judge_init():
    global file_var, video_slots, root

    print("子 UI 初始化")
    root = tk.Tk()
    root.title("動作判斷系統")
    root.geometry("1024x1024")

    tk.Label(root, text="選擇影片檔案：", font=("Arial", 18, "bold")).pack(pady=10)
    file_var = tk.StringVar()
    video_files = get_video_files()
    if video_files:
        file_var.set(video_files[0])
    else:
        file_var.set("（找不到影片）")
        print("No file path")
    tk.OptionMenu(root, file_var, *video_files).pack(pady=5)

    tk.Button(root, text="開始判斷", command=start_action, bg="green", fg="white").pack(pady=20)

    frame_count = 15
    cols = 5
    rows = 3
    frame_index = 0

    main_frame = tk.Frame(root)
    main_frame.pack()

    video_slots = []

    for r in range(rows):
        for c in range(cols):
            if frame_index >= frame_count:
                break

            video_frame = tk.LabelFrame(main_frame, text=f"動作 {frame_index + 1}", padx=5, pady=5)
            video_frame.grid(row=r, column=c, padx=10, pady=10)

            headers = SCORE_TYPES[frame_index]
            slot_labels = []

            for i, header in enumerate(headers):
                label = tk.Label(video_frame, text=header, font=("Arial", 9, "bold"))
                label.grid(row=i, column=0, padx=3, pady=2)

                slot = tk.Label(video_frame, text="", relief="ridge", width=8, height=2)
                slot.grid(row=i, column=1, padx=3, pady=2)
                slot_labels.append(slot)

            video_slots.append(slot_labels)
            frame_index += 1

    root.mainloop()

if __name__ == "__main__":
    judge_init()